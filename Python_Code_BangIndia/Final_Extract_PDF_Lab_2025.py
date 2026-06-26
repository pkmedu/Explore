# -*- coding: utf-8 -*-
"""
Created on Tue Jul 22 12:10:13 2025

@author: muhuri
"""

# -*- coding: utf-8 -*-
"""
Enhanced Lab Data Extractor - Improved Version
Key improvements:
1. Better error handling and logging
2. Configuration management
3. Validation functions
4. Progress tracking
5. More robust pattern matching
"""

import fitz
import pandas as pd
import re
import os
import logging
from datetime import datetime
from typing import List, Dict, Optional, Tuple
from dataclasses import dataclass

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('lab_extraction_2025.log'),
        logging.StreamHandler()
    ]
)

@dataclass
class ExtractionConfig:
    """Configuration class for extraction parameters"""
    flag_symbols: List[str] = None
    target_tests: List[str] = None
    text_results: List[str] = None
    max_lookback_lines: int = 5
    max_lookahead_lines: int = 8
    duplicate_line_threshold: int = 5
    
    def __post_init__(self):
        if self.flag_symbols is None:
            self.flag_symbols = ['!', '▲', '△', '⚠', '★', '*', '•', '◆', '♦', '▼', '⬥', '⚪', '🔺', '❗', '‼']
        
        if self.target_tests is None:
            self.target_tests = [
                'LDL-CHOLESTEROL', 'LDL CHOLESTEROL', 'GLUCOSE', 'MCHC', 
                'HEMOGLOBIN A1C', 'HEMOGLOBIN A1c', 'LDL', 'HbA1c', 'A1C', 
                'HYALINE CAST', 'CHOL/HDLC RATIO'
            ]
        
        if self.text_results is None:
            self.text_results = [
                'NEGATIVE', 'POSITIVE', 'NORMAL', 'ABNORMAL', 
                'CLEAR', 'DETECTED', 'NOT DETECTED', 'PRESENT', 'ABSENT'
            ]

class LabDataExtractor:
    def __init__(self, config: ExtractionConfig = None):
        self.config = config or ExtractionConfig()
        self.logger = logging.getLogger(__name__)
        self.extracted_data = []
        
    def validate_inputs(self, pdf_path: str, excel_path: str) -> bool:
        """Validate input parameters"""
        if not pdf_path or not isinstance(pdf_path, str):
            self.logger.error("PDF path must be a non-empty string")
            return False
        
        if not os.path.exists(pdf_path):
            self.logger.error(f"PDF file not found: {pdf_path}")
            return False
        
        if not excel_path or not isinstance(excel_path, str):
            self.logger.error("Excel path must be a non-empty string")
            return False
        
        # Check if we can write to the output directory
        output_dir = os.path.dirname(excel_path)
        if output_dir and not os.path.exists(output_dir):
            try:
                os.makedirs(output_dir)
                self.logger.info(f"Created output directory: {output_dir}")
            except Exception as e:
                self.logger.error(f"Cannot create output directory: {e}")
                return False
        
        return True
    
    def is_valid_value(self, value_line: str) -> Tuple[bool, str]:
        """Check if a line contains a valid test value and return its type"""
        if not value_line or not value_line.strip():
            return False, None
        
        value_line = value_line.strip()
        
        # Priority order for value types
        patterns = [
            (r'^\d+\.?\d*\s*[HL]$', 'Flagged'),           # H/L flagged values (highest priority)
            (r'^[<>=]\s*\d+\.?\d*$', 'Comparison'),       # <5, >10, =50
            (r'^\d+\s*-\s*\d+$', 'Range'),                # 1-5, 10-20
            (r'^\d+\.?\d*$', 'Numeric'),                  # Plain numbers
        ]
        
        # Check patterns in order
        for pattern, value_type in patterns:
            if re.match(pattern, value_line):
                return True, value_type
        
        # Check text results
        if value_line.upper() in self.config.text_results:
            return True, 'Text'
        
        return False, None
    
    def parse_value(self, value_line: str, value_type: str) -> any:
        """Parse a value based on its type"""
        try:
            if value_type == 'Flagged':
                return value_line  # Keep "102 H" as is
            elif value_type == 'Numeric':
                return float(value_line)
            elif value_type in ['Range', 'Text', 'Comparison']:
                return value_line
            else:
                return value_line
        except (ValueError, TypeError):
            self.logger.warning(f"Could not parse value: {value_line} as type {value_type}")
            return value_line
    
    def extract_by_reference_range(self, lines: List[str], page_num: int) -> List[Dict]:
        """Extract tests using reference range pattern"""
        results = []
        self.logger.info("Using reference range method...")
        
        for i, line in enumerate(lines):
            line = line.strip()
            
            if "Reference Range:" not in line:
                continue
            
            self.logger.debug(f"Found reference range: {line}")
            
            # Look backwards for value
            value, value_type, value_line_idx = None, None, None
            
            for j in range(i-1, max(i-self.config.max_lookback_lines, -1), -1):
                if j < 0:
                    break
                
                prev_line = lines[j].strip()
                is_valid, val_type = self.is_valid_value(prev_line)
                
                if is_valid:
                    value = self.parse_value(prev_line, val_type)
                    value_type = val_type
                    value_line_idx = j
                    self.logger.debug(f"Found value: {value} ({value_type})")
                    break
            
            if value is None or value_line_idx is None:
                continue
            
            # Look backwards for test name
            test_name, test_name_line_idx = None, None
            
            for k in range(value_line_idx-1, max(value_line_idx-self.config.max_lookback_lines, -1), -1):
                if k < 0:
                    break
                
                test_line = lines[k].strip()
                
                # Skip invalid test name patterns
                if (not test_line or len(test_line) < 3 or 
                    test_line.startswith('Reference Range') or
                    self.is_valid_value(test_line)[0]):
                    continue
                
                test_name = test_line
                test_name_line_idx = k
                self.logger.debug(f"Found test name: {test_name}")
                break
            
            if test_name and test_name_line_idx is not None:
                ref_range = line.replace("Reference Range:", "").strip()
                
                result = {
                    'Test_Name': test_name,
                    'Value': value,
                    'Value_Type': value_type,
                    'Reference_Range': ref_range,
                    'Page': page_num,
                    'Line_Position': test_name_line_idx,
                    'Has_Triangle_Symbol': self.is_target_test(test_name),
                    'Method': 'reference_range'
                }
                
                results.append(result)
                triangle_note = " [TARGET]" if result['Has_Triangle_Symbol'] else ""
                self.logger.info(f"Extracted: {test_name} = {value}{triangle_note}")
        
        return results
    
    def extract_by_flag_symbols(self, lines: List[str], page_num: int) -> List[Dict]:
        """Extract tests that start with flag symbols"""
        results = []
        self.logger.info("Searching for flag symbol tests...")
        
        for i, line in enumerate(lines):
            line = line.strip()
            
            # Check if line starts with any flag symbol
            starts_with_flag = any(line.startswith(symbol) for symbol in self.config.flag_symbols)
            
            if not starts_with_flag or len(line) <= 2:
                continue
            
            self.logger.debug(f"Found flag symbol line: {line}")
            
            # Look ahead for value
            for j in range(i+1, min(i+self.config.max_lookahead_lines, len(lines))):
                value_line = lines[j].strip()
                
                if not value_line:
                    continue
                
                is_valid, value_type = self.is_valid_value(value_line)
                
                if is_valid:
                    value = self.parse_value(value_line, value_type)
                    
                    # Look for reference range
                    ref_range = ""
                    for k in range(j+1, min(j+4, len(lines))):
                        if k < len(lines) and "Reference Range:" in lines[k]:
                            ref_range = lines[k].replace("Reference Range:", "").strip()
                            break
                    
                    if value and str(value).strip():
                        result = {
                            'Test_Name': line,
                            'Value': value,
                            'Value_Type': value_type,
                            'Reference_Range': ref_range or None,
                            'Page': page_num,
                            'Line_Position': i,
                            'Has_Triangle_Symbol': True,
                            'Method': 'flag_symbol'
                        }
                        
                        results.append(result)
                        self.logger.info(f"Captured flag symbol test: {line} = {value}")
                        break
        
        return results
    
    def extract_by_target_names(self, lines: List[str], page_num: int) -> List[Dict]:
        """Extract specific target test names"""
        results = []
        self.logger.info("Searching for target test names...")
        
        for i, line in enumerate(lines):
            line = line.strip()
            
            # Check if this line matches any target test
            matched_target = None
            for target_name in self.config.target_tests:
                if self.normalize_test_name(line) == self.normalize_test_name(target_name):
                    matched_target = target_name
                    break
            
            if not matched_target:
                continue
            
            self.logger.debug(f"Found target test: '{line}' matches '{matched_target}'")
            
            # Skip if already captured by checking existing results
            if self.is_already_captured(line, page_num, i):
                self.logger.debug("Already captured, skipping")
                continue
            
            # Look ahead for value
            for j in range(i+1, min(i+self.config.max_lookahead_lines, len(lines))):
                value_line = lines[j].strip()
                
                if not value_line:
                    continue
                
                is_valid, value_type = self.is_valid_value(value_line)
                
                if is_valid:
                    value = self.parse_value(value_line, value_type)
                    
                    # Look for reference range
                    ref_range = ""
                    for k in range(j+1, min(j+4, len(lines))):
                        if k < len(lines) and ("Reference Range:" in lines[k] or "Reference range:" in lines[k]):
                            ref_range = lines[k].replace("Reference Range:", "").replace("Reference range:", "").strip()
                            break
                    
                    result = {
                        'Test_Name': line,
                        'Value': value,
                        'Value_Type': value_type,
                        'Reference_Range': ref_range or None,
                        'Page': page_num,
                        'Line_Position': i,
                        'Has_Triangle_Symbol': self.is_target_test(line),
                        'Method': 'target_name_search'
                    }
                    
                    results.append(result)
                    triangle_note = " [TARGET]" if result['Has_Triangle_Symbol'] else ""
                    self.logger.info(f"Captured target: {line} = {value}{triangle_note}")
                    break
        
        return results
    
    def normalize_test_name(self, test_name: str) -> str:
        """Normalize test name for comparison"""
        return test_name.upper().replace('-', ' ').replace(' ', '')
    
    def is_target_test(self, test_name: str) -> bool:
        """Check if a test is one of our target tests"""
        target_names = [
            'LDL-CHOLESTEROL', 'GLUCOSE', 'MCHC', 'HEMOGLOBIN A1C', 
            'HEMOGLOBIN A1c', 'HYALINE CAST', 'CHOL/HDLC RATIO'
        ]
        normalized_test = self.normalize_test_name(test_name)
        return any(self.normalize_test_name(target) == normalized_test for target in target_names)
    
    def is_already_captured(self, test_name: str, page_num: int, line_pos: int) -> bool:
        """Check if a test has already been captured"""
        for existing in self.extracted_data:
            if (existing['Test_Name'] == test_name and 
                existing['Page'] == page_num and 
                abs(existing['Line_Position'] - line_pos) <= 2):
                return True
        return False
    
    def remove_duplicates(self, data: List[Dict]) -> List[Dict]:
        """Remove duplicate entries while preserving order"""
        df = pd.DataFrame(data)
        df_sorted = df.sort_values(['Page', 'Line_Position']).reset_index(drop=True)
        
        filtered_data = []
        seen_tests = {}
        
        for _, row in df_sorted.iterrows():
            test_key = f"{row['Test_Name'].upper()}_{row['Page']}"
            current_line = row['Line_Position']
            current_value = str(row['Value']) if row['Value'] is not None else ""
            
            is_duplicate = False
            
            if test_key in seen_tests:
                prev_line = seen_tests[test_key]['line']
                prev_value = seen_tests[test_key]['value']
                line_distance = abs(current_line - prev_line)
                
                if (line_distance <= self.config.duplicate_line_threshold and 
                    (current_value == prev_value or 
                     (not current_value.strip() and not prev_value.strip()))):
                    is_duplicate = True
                    self.logger.debug(f"Skipped duplicate: {row['Test_Name']} (Line {current_line})")
                else:
                    seen_tests[test_key] = {'line': current_line, 'value': current_value}
            else:
                seen_tests[test_key] = {'line': current_line, 'value': current_value}
            
            if not is_duplicate:
                filtered_data.append(row.to_dict())
        
        return filtered_data
    
    def apply_excel_formatting(self, excel_path: str, df: pd.DataFrame):
        """Apply formatting to the Excel file"""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, Alignment
            
            wb = load_workbook(excel_path)
            ws = wb.active
            
            red_font = Font(color='FF0000', bold=True)
            right_align = Alignment(horizontal='right')
            
            # Find column indices
            value_col_idx = None
            ref_range_col_idx = None
            
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value == 'Value':
                    value_col_idx = col_idx
                elif cell.value == 'Reference_Range':
                    ref_range_col_idx = col_idx
            
            # Format Value column
            if value_col_idx:
                flagged_count = 0
                chol_formatted = 0
                
                for row_idx in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=value_col_idx)
                    cell.alignment = right_align
                    
                    # Get test name
                    test_name_cell = ws.cell(row=row_idx, column=1)
                    test_name = str(test_name_cell.value) if test_name_cell.value else ""
                    
                    # Special formatting for CHOL/HDLC RATIO
                    if "CHOL/HDLC" in test_name.upper() and cell.value is not None:
                        try:
                            original_str = str(cell.value)
                            numeric_value = float(original_str.replace(' H', '').replace(' L', '').strip())
                            
                            has_h_flag = 'H' in original_str
                            has_l_flag = 'L' in original_str
                            
                            formatted_value = f"{numeric_value:.1f}"
                            if has_h_flag:
                                formatted_value += " H"
                            elif has_l_flag:
                                formatted_value += " L"
                            
                            cell.value = formatted_value
                            chol_formatted += 1
                            
                            if has_h_flag or has_l_flag:
                                cell.font = red_font
                                flagged_count += 1
                        except (ValueError, TypeError):
                            pass
                    
                    # Regular flagged value formatting
                    elif cell.value and isinstance(cell.value, str):
                        if any(flag in str(cell.value) for flag in [' H', ' L', 'H', 'L']):
                            cell.font = red_font
                            flagged_count += 1
                
                self.logger.info(f"Applied formatting to {flagged_count} flagged values and {chol_formatted} cholesterol ratios")
            
            # Format Reference Range column
            if ref_range_col_idx:
                for row_idx in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=ref_range_col_idx)
                    if cell.value:
                        cell.alignment = right_align
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                if adjusted_width < 12:
                    adjusted_width = 12
                
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(excel_path)
            self.logger.info("Excel formatting applied successfully")
            
        except ImportError:
            self.logger.warning("openpyxl not available. Install with: pip install openpyxl")
        except Exception as e:
            self.logger.error(f"Could not apply formatting: {e}")
    
    def extract_data(self, pdf_path: str, excel_path: str) -> Optional[pd.DataFrame]:
        """Main extraction method"""
        if not self.validate_inputs(pdf_path, excel_path):
            return None
        
        self.logger.info(f"Starting enhanced extraction from: {pdf_path}")
        self.extracted_data = []
        
        try:
            doc = fitz.open(pdf_path)
            total_pages = len(doc)
            
            for page_num, page in enumerate(doc, 1):
                self.logger.info(f"Processing page {page_num}/{total_pages}...")
                
                try:
                    text = page.get_text()
                    lines = text.split('\n')
                    
                    # Apply all extraction methods
                    page_results = []
                    page_results.extend(self.extract_by_reference_range(lines, page_num))
                    page_results.extend(self.extract_by_flag_symbols(lines, page_num))
                    page_results.extend(self.extract_by_target_names(lines, page_num))
                    
                    self.extracted_data.extend(page_results)
                    
                except Exception as e:
                    self.logger.error(f"Error processing page {page_num}: {e}")
                    continue
            
            doc.close()
            
            if not self.extracted_data:
                self.logger.warning("No data extracted")
                return None
            
            # Remove duplicates and create DataFrame
            filtered_data = self.remove_duplicates(self.extracted_data)
            df = pd.DataFrame(filtered_data)
            
            # Filter out missing values for non-target tests
            df_final = df[
                (df['Value'].notna() & (df['Value'].astype(str).str.strip() != '')) |
                (df['Has_Triangle_Symbol'] == True)
            ].copy()
            
            # Add extraction timestamp
            df_final['Extraction_Date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # Select output columns
            output_columns = ['Test_Name', 'Value', 'Reference_Range', 'Extraction_Date']
            df_output = df_final[output_columns].copy()
            
            # Save to Excel
            df_output.to_excel(excel_path, index=False)
            
            # Apply formatting
            self.apply_excel_formatting(excel_path, df_output)
            
            # Log summary
            self.log_summary(df_final, excel_path)
            
            return df_final
            
        except Exception as e:
            self.logger.error(f"Extraction failed: {e}")
            return None
    
    def log_summary(self, df: pd.DataFrame, excel_path: str):
        """Log extraction summary"""
        total_tests = len(df)
        target_tests = df['Has_Triangle_Symbol'].sum()
        
        self.logger.info(f"Extraction completed successfully!")
        self.logger.info(f"Total tests extracted: {total_tests}")
        self.logger.info(f"Target tests found: {target_tests}")
        self.logger.info(f"Output saved to: {excel_path}")
        
        if target_tests > 0:
            target_found = df[df['Has_Triangle_Symbol'] == True]
            self.logger.info("Target tests with triangle symbols:")
            for _, row in target_found.iterrows():
                value_display = row['Value'] if row['Value'] is not None else "MISSING"
                self.logger.info(f"  - {row['Test_Name']}: {value_display}")

# Example usage
def main():
    """Main execution function"""
    config = ExtractionConfig()
    extractor = LabDataExtractor(config)
    
    pdf_file = r'C:\Data\LabReport_2025.pdf'
    excel_file = r'C:\Data2\LabReport_2025_enhanced.xlsx'
    
    result = extractor.extract_data(pdf_file, excel_file)
    
    if result is not None:
        print("Extraction completed successfully!")
        print(f"Check the log file 'lab_extraction.log' for detailed information")
    else:
        print("Extraction failed. Check the log file for error details.")

if __name__ == "__main__":
    main()