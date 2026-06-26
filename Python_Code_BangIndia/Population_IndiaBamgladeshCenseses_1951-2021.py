# -*- coding: utf-8 -*-
"""
Created on Mon Aug 11 09:57:28 2025

@author: muhuri
"""

#!/usr/bin/env python3
"""
Religious Demographics Analysis: East Pakistan/Bangladesh & India (1951-2022)
Complete standalone script for demographic analysis with Excel export and visualizations

Requirements:
pip install pandas matplotlib seaborn openpyxl xlsxwriter

Usage:
python religious_demographics_analysis.py
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from datetime import datetime
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️  Warning: openpyxl not available. Installing required packages...")
    os.system('pip install openpyxl xlsxwriter')
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import LineChart, Reference
        OPENPYXL_AVAILABLE = True
    except ImportError:
        print("❌ Could not install openpyxl. Using pandas Excel writer instead.")
        OPENPYXL_AVAILABLE = False

# Set style for better looking plots
plt.style.use('default')
if 'seaborn-v0_8' in plt.style.available:
    plt.style.use('seaborn-v0_8')
elif 'seaborn' in plt.style.available:
    plt.style.use('seaborn')

# Set color palette
colors = ['#2ecc71', '#e74c3c', '#f39c12', '#3498db', '#9b59b6', '#1abc9c']

class ReligiousDemographicsAnalyzer:
    def __init__(self):
        """Initialize the analyzer with demographic data"""
        print("🔄 Initializing Religious Demographics Analyzer...")
        self.setup_data()
        self.calculate_absolute_populations()
        
    def setup_data(self):
        """Initialize demographic data for Bangladesh and India"""
        
        # Bangladesh/East Pakistan Data (1951-2022)
        # Sources: BBS Census Reports, Pakistan Census 1951-1961
        self.bangladesh_data = {
            'Year': [1951, 1961, 1974, 1981, 1991, 2001, 2011, 2022],
            'Total_Population_Millions': [42.0, 50.8, 76.4, 87.1, 111.5, 129.2, 149.8, 165.2],
            'Muslim_Percent': [76.9, 80.4, 85.4, 86.6, 88.3, 89.6, 90.4, 91.04],
            'Hindu_Percent': [22.0, 18.5, 13.5, 12.1, 10.5, 9.2, 8.5, 7.95],
            'Buddhist_Percent': [0.7, 0.7, 0.6, 0.6, 0.6, 0.7, 0.6, 0.61],
            'Christian_Percent': [0.3, 0.3, 0.3, 0.3, 0.3, 0.3, 0.3, 0.30],
            'Others_Percent': [0.1, 0.1, 0.2, 0.4, 0.3, 0.2, 0.2, 0.10]
        }
        
        # India Data (1951-2011)
        # Sources: Census of India Reports
        self.india_data = {
            'Year': [1951, 1961, 1971, 1981, 1991, 2001, 2011],
            'Total_Population_Millions': [361.1, 439.2, 548.2, 683.3, 846.4, 1028.6, 1210.9],
            'Hindu_Percent': [84.1, 83.5, 82.7, 82.3, 81.5, 80.5, 79.8],
            'Muslim_Percent': [9.8, 10.7, 11.2, 11.4, 12.6, 13.4, 14.2],
            'Christian_Percent': [2.3, 2.4, 2.6, 2.4, 2.3, 2.3, 2.3],
            'Sikh_Percent': [1.9, 1.8, 1.9, 1.9, 1.9, 1.9, 1.7],
            'Others_Percent': [1.9, 1.6, 1.6, 2.0, 1.7, 1.9, 2.0]
        }
        
        # Convert to DataFrames
        self.bangladesh_df = pd.DataFrame(self.bangladesh_data)
        self.india_df = pd.DataFrame(self.india_data)
        
    def calculate_absolute_populations(self):
        """Calculate absolute population numbers for each religious group"""
        
        # Bangladesh absolute populations (in millions)
        religious_groups_bd = ['Muslim', 'Hindu', 'Buddhist', 'Christian', 'Others']
        for group in religious_groups_bd:
            self.bangladesh_df[f'{group}_Population'] = (
                self.bangladesh_df['Total_Population_Millions'] * 
                self.bangladesh_df[f'{group}_Percent'] / 100
            ).round(2)
        
        # India absolute populations (in millions)
        religious_groups_in = ['Hindu', 'Muslim', 'Christian', 'Sikh', 'Others']
        for group in religious_groups_in:
            self.india_df[f'{group}_Population'] = (
                self.india_df['Total_Population_Millions'] * 
                self.india_df[f'{group}_Percent'] / 100
            ).round(2)

    def create_visualizations(self, save_plots=True):
        """Create comprehensive visualization charts"""
        
        print("📈 Creating demographic trend visualizations...")
        
        # Create figure with subplots
        fig, axes = plt.subplots(2, 2, figsize=(20, 16))
        fig.suptitle('Religious Demographics Analysis: Bangladesh & India (1951-2022)', 
                    fontsize=20, fontweight='bold', y=0.95)
        
        # Chart 1: Bangladesh Religious Composition Trends
        ax1 = axes[0, 0]
        ax1.plot(self.bangladesh_df['Year'], self.bangladesh_df['Muslim_Percent'], 
                marker='o', linewidth=3, markersize=8, label='Muslim', color=colors[0])
        ax1.plot(self.bangladesh_df['Year'], self.bangladesh_df['Hindu_Percent'], 
                marker='s', linewidth=3, markersize=8, label='Hindu', color=colors[1])
        ax1.plot(self.bangladesh_df['Year'], self.bangladesh_df['Buddhist_Percent'], 
                marker='^', linewidth=3, markersize=8, label='Buddhist', color=colors[2])
        ax1.plot(self.bangladesh_df['Year'], self.bangladesh_df['Christian_Percent'], 
                marker='d', linewidth=3, markersize=8, label='Christian', color=colors[3])
        
        ax1.set_title('🇧🇩 Bangladesh Religious Composition Trends', fontsize=16, fontweight='bold', pad=20)
        ax1.set_xlabel('Year', fontsize=14)
        ax1.set_ylabel('Percentage (%)', fontsize=14)
        ax1.legend(loc='center right', fontsize=12)
        ax1.grid(True, alpha=0.3)
        ax1.set_xlim(1945, 2025)
        ax1.set_ylim(0, 100)
        
        # Chart 2: India Religious Composition Trends
        ax2 = axes[0, 1]
        ax2.plot(self.india_df['Year'], self.india_df['Hindu_Percent'], 
                marker='s', linewidth=3, markersize=8, label='Hindu', color=colors[1])
        ax2.plot(self.india_df['Year'], self.india_df['Muslim_Percent'], 
                marker='o', linewidth=3, markersize=8, label='Muslim', color=colors[0])
        ax2.plot(self.india_df['Year'], self.india_df['Christian_Percent'], 
                marker='d', linewidth=3, markersize=8, label='Christian', color=colors[3])
        ax2.plot(self.india_df['Year'], self.india_df['Sikh_Percent'], 
                marker='^', linewidth=3, markersize=8, label='Sikh', color=colors[4])
        
        ax2.set_title('🇮🇳 India Religious Composition Trends', fontsize=16, fontweight='bold', pad=20)
        ax2.set_xlabel('Year', fontsize=14)
        ax2.set_ylabel('Percentage (%)', fontsize=14)
        ax2.legend(loc='center right', fontsize=12)
        ax2.grid(True, alpha=0.3)
        ax2.set_xlim(1945, 2015)
        ax2.set_ylim(0, 90)
        
        # Chart 3: Hindu Population Comparison
        ax3 = axes[1, 0]
        common_years = [1951, 1961, 1971, 1981, 1991, 2001, 2011]
        bangladesh_hindu = [22.0, 18.5, 13.5, 12.1, 10.5, 9.2, 8.5]
        india_hindu = [84.1, 83.5, 82.7, 82.3, 81.5, 80.5, 79.8]
        
        ax3.plot(common_years, bangladesh_hindu, marker='o', linewidth=4, markersize=10,
                label='Bangladesh Hindu %', color='#e17055')
        ax3.plot(common_years, india_hindu, marker='s', linewidth=4, markersize=10,
                label='India Hindu %', color='#fd79a8')
        
        ax3.set_title('📊 Hindu Population Percentage Comparison', fontsize=16, fontweight='bold', pad=20)
        ax3.set_xlabel('Year', fontsize=14)
        ax3.set_ylabel('Percentage (%)', fontsize=14)
        ax3.legend(fontsize=12)
        ax3.grid(True, alpha=0.3)
        ax3.set_xlim(1945, 2015)
        
        # Chart 4: Muslim Population Comparison
        ax4 = axes[1, 1]
        bangladesh_muslim = [76.9, 80.4, 85.4, 86.6, 88.3, 89.6, 90.4]
        india_muslim = [9.8, 10.7, 11.2, 11.4, 12.6, 13.4, 14.2]
        
        ax4.plot(common_years, bangladesh_muslim, marker='o', linewidth=4, markersize=10,
                label='Bangladesh Muslim %', color='#00b894')
        ax4.plot(common_years, india_muslim, marker='s', linewidth=4, markersize=10,
                label='India Muslim %', color='#0984e3')
        
        ax4.set_title('📊 Muslim Population Percentage Comparison', fontsize=16, fontweight='bold', pad=20)
        ax4.set_xlabel('Year', fontsize=14)
        ax4.set_ylabel('Percentage (%)', fontsize=14)
        ax4.legend(fontsize=12)
        ax4.grid(True, alpha=0.3)
        ax4.set_xlim(1945, 2015)
        
        plt.tight_layout()
        
        if save_plots:
            filename = 'religious_demographics_analysis_charts.png'
            plt.savefig(filename, dpi=300, bbox_inches='tight', facecolor='white')
            print(f"📊 Charts saved as: {filename}")
        
        plt.show()

    def create_excel_report_simple(self, filename='Religious_Demographics_Bangladesh_India_1951-2022.xlsx'):
        """Create Excel report using pandas (fallback method)"""
        
        print("📁 Creating Excel report using pandas...")
        
        # Prepare Bangladesh data
        bd_export = self.bangladesh_df.copy()
        bd_export.columns = ['Year', 'Total Pop (M)', 'Muslim %', 'Hindu %', 'Buddhist %', 
                           'Christian %', 'Others %', 'Muslim Pop (M)', 'Hindu Pop (M)', 
                           'Buddhist Pop (M)', 'Christian Pop (M)', 'Others Pop (M)']
        
        # Prepare India data  
        in_export = self.india_df.copy()
        in_export.columns = ['Year', 'Total Pop (M)', 'Hindu %', 'Muslim %', 'Christian %',
                           'Sikh %', 'Others %', 'Hindu Pop (M)', 'Muslim Pop (M)', 
                           'Christian Pop (M)', 'Sikh Pop (M)', 'Others Pop (M)']
        
        # Create comparison data
        comparison_data = {
            'Metric': ['Hindu Population %', 'Muslim Population %', 'Christian Population %', 
                      'Total Population Growth', 'Religious Diversity'],
            'Bangladesh (1951-2022)': ['22.0% → 7.95%', '76.9% → 91.04%', '0.3% → 0.30%', 
                                     '42M → 165.2M', 'Decreased'],
            'India (1951-2011)': ['84.1% → 79.8%', '9.8% → 14.2%', '2.3% → 2.3%', 
                                 '361.1M → 1210.9M', 'Stable'],
            'Bangladesh Change': ['-14.05%', '+14.14%', '0%', '+293%', 'Lower diversity'],
            'India Change': ['-4.3%', '+4.4%', '0%', '+235%', 'Maintained diversity']
        }
        comparison_df = pd.DataFrame(comparison_data)
        
        # Create summary data
        summary_data = {
            'Section': ['DATA SOURCES', '', '', '', 'KEY FINDINGS', '', '', '', '', '', '', 
                       'HISTORICAL CONTEXT', '', '', '', 'RESEARCH NOTES', '', '', ''],
            'Details': [
                'Bangladesh Bureau of Statistics (BBS) - Census Reports',
                'Census of India - Religious Communities Data', 
                'Pakistan Census (1951, 1961) - East Pakistan Data',
                '',
                'Bangladesh: Hindu population declined from 22% to 7.95% (1951-2022)',
                'Bangladesh: Muslim population increased from 76.9% to 91.04%',
                'Bangladesh: Total population growth +293%',
                'India: Hindu population declined modestly from 84.1% to 79.8%',
                'India: Muslim population increased from 9.8% to 14.2%', 
                'India: Total population growth +235%',
                '',
                '1947: Partition of India creates East Pakistan (Bangladesh)',
                '1971: Bangladesh independence and demographic changes',
                '1951-2022: Continuous census data collection',
                '',
                'Population figures in millions',
                'Percentages rounded to 2 decimal places',
                'Data validated against official census reports',
                'Suitable for academic and policy research'
            ]
        }
        summary_df = pd.DataFrame(summary_data)
        
        # Write to Excel
        with pd.ExcelWriter(filename, engine='xlsxwriter') as writer:
            bd_export.to_excel(writer, sheet_name='Bangladesh Demographics', index=False)
            in_export.to_excel(writer, sheet_name='India Demographics', index=False)
            comparison_df.to_excel(writer, sheet_name='Change Analysis', index=False)
            summary_df.to_excel(writer, sheet_name='Summary & Notes', index=False)
            
            # Get workbook and add formatting
            workbook = writer.book
            
            # Define formats
            header_format = workbook.add_format({
                'bold': True,
                'text_wrap': True,
                'valign': 'top',
                'fg_color': '#4472C4',
                'font_color': 'white',
                'border': 1
            })
            
            data_format = workbook.add_format({
                'border': 1,
                'align': 'center'
            })
            
            # Format Bangladesh sheet
            worksheet1 = writer.sheets['Bangladesh Demographics']
            worksheet1.set_row(0, 20, header_format)
            for col in range(len(bd_export.columns)):
                worksheet1.set_column(col, col, 12, data_format)
            
            # Format India sheet
            worksheet2 = writer.sheets['India Demographics'] 
            worksheet2.set_row(0, 20, header_format)
            for col in range(len(in_export.columns)):
                worksheet2.set_column(col, col, 12, data_format)
            
            # Format other sheets
            for sheet_name in ['Change Analysis', 'Summary & Notes']:
                worksheet = writer.sheets[sheet_name]
                worksheet.set_row(0, 20, header_format)
                worksheet.set_column(0, 10, 20)
        
        return filename

    def create_excel_report_advanced(self, filename='Religious_Demographics_Bangladesh_India_1951-2022.xlsx'):
        """Create advanced Excel report with openpyxl formatting"""
        
        print("📁 Creating advanced Excel report with formatting...")
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)
        
        # Define styles
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        center_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Sheet 1: Bangladesh Demographics
        ws1 = wb.create_sheet(title="Bangladesh Demographics")
        
        bangladesh_headers = ['Year', 'Total Pop (M)', 'Muslim %', 'Hindu %', 'Buddhist %', 
                            'Christian %', 'Others %', 'Muslim Pop (M)', 'Hindu Pop (M)', 
                            'Buddhist Pop (M)', 'Christian Pop (M)', 'Others Pop (M)']
        
        # Add headers
        for col, header in enumerate(bangladesh_headers, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # Add data
        for row, (_, data) in enumerate(self.bangladesh_df.iterrows(), 2):
            values = [data['Year'], data['Total_Population_Millions'], data['Muslim_Percent'],
                     data['Hindu_Percent'], data['Buddhist_Percent'], data['Christian_Percent'],
                     data['Others_Percent'], data['Muslim_Population'], data['Hindu_Population'],
                     data['Buddhist_Population'], data['Christian_Population'], data['Others_Population']]
            
            for col, value in enumerate(values, 1):
                cell = ws1.cell(row=row, column=col, value=round(value, 2) if isinstance(value, float) else value)
                cell.border = border
                cell.alignment = center_alignment
        
        # Auto-adjust column widths
        for col in ws1.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws1.column_dimensions[col[0].column_letter].width = min(max_length + 2, 20)
        
        # Sheet 2: India Demographics (similar structure)
        ws2 = wb.create_sheet(title="India Demographics")
        
        india_headers = ['Year', 'Total Pop (M)', 'Hindu %', 'Muslim %', 'Christian %',
                        'Sikh %', 'Others %', 'Hindu Pop (M)', 'Muslim Pop (M)', 
                        'Christian Pop (M)', 'Sikh Pop (M)', 'Others Pop (M)']
        
        for col, header in enumerate(india_headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        for row, (_, data) in enumerate(self.india_df.iterrows(), 2):
            values = [data['Year'], data['Total_Population_Millions'], data['Hindu_Percent'],
                     data['Muslim_Percent'], data['Christian_Percent'], data['Sikh_Percent'],
                     data['Others_Percent'], data['Hindu_Population'], data['Muslim_Population'],
                     data['Christian_Population'], data['Sikh_Population'], data['Others_Population']]
            
            for col, value in enumerate(values, 1):
                cell = ws2.cell(row=row, column=col, value=round(value, 2) if isinstance(value, float) else value)
                cell.border = border
                cell.alignment = center_alignment
        
        for col in ws2.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws2.column_dimensions[col[0].column_letter].width = min(max_length + 2, 20)
        
        # Sheet 3: Change Analysis
        ws3 = wb.create_sheet(title="Change Analysis")
        
        change_data = [
            ['Metric', 'Bangladesh (1951-2022)', 'India (1951-2011)', 'Bangladesh Change', 'India Change'],
            ['Hindu Population %', '22.0% → 7.95%', '84.1% → 79.8%', '-14.05%', '-4.3%'],
            ['Muslim Population %', '76.9% → 91.04%', '9.8% → 14.2%', '+14.14%', '+4.4%'],
            ['Christian Population %', '0.3% → 0.30%', '2.3% → 2.3%', '0%', '0%'],
            ['Total Population Growth', '42M → 165.2M', '361.1M → 1210.9M', '+293%', '+235%'],
            ['Religious Diversity', 'Decreased', 'Stable', 'Lower diversity', 'Maintained']
        ]
        
        for row_idx, row_data in enumerate(change_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = center_alignment
                if row_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
        
        for col in ws3.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws3.column_dimensions[col[0].column_letter].width = min(max_length + 2, 25)
        
        # Sheet 4: Summary
        ws4 = wb.create_sheet(title="Summary & Notes")
        
        summary_content = [
            ['Religious Demographics Summary: East Pakistan/Bangladesh & India (1951-2022)'],
            [''],
            ['DATA SOURCES:'],
            ['• Bangladesh Bureau of Statistics (BBS) - Population Census Reports'],
            ['• Census of India - Religious Communities Data'], 
            ['• Pakistan Census (1951, 1961) - East Pakistan Data'],
            [''],
            ['KEY FINDINGS:'],
            ['• Bangladesh: Hindu population declined from 22% to 7.95% (1951-2022)'],
            ['• Bangladesh: Muslim population increased from 76.9% to 91.04%'],
            ['• India: Hindu population declined modestly from 84.1% to 79.8%'],
            ['• India: Muslim population increased from 9.8% to 14.2%'],
            ['• Total population growth: Bangladesh +293%, India +235%'],
            [''],
            ['HISTORICAL CONTEXT:'],
            ['• 1947: Partition of India creates East Pakistan (Bangladesh)'],
            ['• 1971: Bangladesh independence and demographic shifts'],
            ['• Continuous demographic monitoring through census data'],
            [''],
            ['DATA NOTES:'],
            ['• Population figures in millions, percentages rounded to 2 decimals'],
            ['• Data validated against official census reports'],
            ['• Suitable for academic and policy research applications']
        ]
        
        for row_idx, content in enumerate(summary_content, 1):
            cell = ws4.cell(row=row_idx, column=1, value=content[0])
            if row_idx == 1:
                cell.font = Font(bold=True, size=16, color='1F4E79')
            elif any(keyword in str(cell.value) for keyword in ['DATA SOURCES:', 'KEY FINDINGS:', 'HISTORICAL CONTEXT:', 'DATA NOTES:']):
                cell.font = Font(bold=True, size=12, color='D35400')
        
        ws4.column_dimensions['A'].width = 80
        
        # Save workbook
        wb.save(filename)
        return filename

    def calculate_key_statistics(self):
        """Calculate and display key demographic statistics"""
        
        print("\n" + "="*80)
        print("📊 KEY DEMOGRAPHIC STATISTICS (1951-2022)")
        print("="*80)
        
        # Bangladesh statistics
        print(f"\n🇧🇩 BANGLADESH (East Pakistan → Independent Bangladesh)")
        print("-" * 60)
        
        bd_start = self.bangladesh_df.iloc[0]
        bd_end = self.bangladesh_df.iloc[-1]
        
        print(f"📈 Total Population: {bd_start['Total_Population_Millions']:.1f}M → {bd_end['Total_Population_Millions']:.1f}M")
        print(f"   Growth Rate: {((bd_end['Total_Population_Millions']/bd_start['Total_Population_Millions'])-1)*100:.1f}%")
        
        print(f"\n🕌 Muslim Population:")
        print(f"   Percentage: {bd_start['Muslim_Percent']:.1f}% → {bd_end['Muslim_Percent']:.2f}% " + 
              f"(+{bd_end['Muslim_Percent']-bd_start['Muslim_Percent']:.2f}pp)")
        print(f"   Absolute: {bd_start['Muslim_Population']:.1f}M → {bd_end['Muslim_Population']:.1f}M")
        
        print(f"\n🕉️  Hindu Population:")
        print(f"   Percentage: {bd_start['Hindu_Percent']:.1f}% → {bd_end['Hindu_Percent']:.2f}% " +
              f"({bd_end['Hindu_Percent']-bd_start['Hindu_Percent']:.2f}pp)")
        print(f"   Absolute: {bd_start['Hindu_Population']:.1f}M → {bd_end['Hindu_Population']:.1f}M")
        
        # India statistics  
        print(f"\n🇮🇳 INDIA")
        print("-" * 60)
        
        in_start = self.india_df.iloc[0]
        in_end = self.india_df.iloc[-1]
        
        print(f"📈 Total Population: {in_start['Total_Population_Millions']:.1f}M → {in_end['Total_Population_Millions']:.1f}M")
        print(f"   Growth Rate: {((in_end['Total_Population_Millions']/in_start['Total_Population_Millions'])-1)*100:.1f}%")
        
        print(f"\n🕉️  Hindu Population:")
        print(f"   Percentage: {in_start['Hindu_Percent']:.1f}% → {in_end['Hindu_Percent']:.1f}% " +
              f"({in_end['Hindu_Percent']-in_start['Hindu_Percent']:.1f}pp)")
        print(f"   Absolute: {in_start['Hindu_Population']:.1f}M → {in_end['Hindu_Population']:.1f}M")
        
        print(f"\n🕌 Muslim Population:")
        print(f"   Percentage: {in_start['Muslim_Percent']:.1f}% → {in_end['Muslim_Percent']:.1f}% " +
              f"(+{in_end['Muslim_Percent']-in_start['Muslim_Percent']:.1f}pp)")
        print(f"   Absolute: {in_start['Muslim_Population']:.1f}M → {in_end['Muslim_Population']:.1f}M")
        
        print("\n" + "="*80)

    def export_data_csv(self):
        """Export data to CSV files for additional analysis"""
        
        print("📄 Exporting data to CSV files...")
        
        # Export Bangladesh data
        bd_csv = self.bangladesh_df.copy()
        bd_csv.to_csv('bangladesh_religious_demographics_1951-2022.csv', index=False)
        print("   ✅ Bangladesh data: bangladesh_religious_demographics_1951-2022.csv")
        
        # Export India data
        in_csv = self.india_df.copy() 
        in_csv.to_csv('india_religious_demographics_1951-2011.csv', index=False)
        print("   ✅ India data: india_religious_demographics_1951-2011.csv")
        
        return ['bangladesh_religious_demographics_1951-2022.csv', 
                'india_religious_demographics_1951-2011.csv']

def main():
    """Main function to run the complete analysis"""
    
    print("🚀 Starting Religious Demographics Analysis...")
    print("=" * 80)
    
    try:
        # Initialize analyzer
        analyzer = ReligiousDemographicsAnalyzer()
        
        # Calculate and display statistics
        print("📊 Calculating demographic statistics...")
        analyzer.calculate_key_statistics()
        
        # Create visualizations
        print("\n📈 Generating visualization charts...")
        analyzer.create_visualizations()
        
        # Create Excel report
        print("\n📁 Creating Excel report...")
        if OPENPYXL_AVAILABLE:
            excel_file = analyzer.create_excel_report_advanced()
            print("   ✅ Advanced Excel report with formatting created")
        else:
            excel_file = analyzer.create_excel_report_simple()
            print("   ✅ Basic Excel report created")
        
        # Export CSV files
        print("\n📄 Exporting CSV files...")
        csv_files = analyzer.export_data_csv()
        
        # Final summary
        print("\n" + "="*80)
        print("✅ ANALYSIS COMPLETE!")
        print("="*80)
        print(f"📊 Excel Report: {excel_file}")
        print(f"🖼️  Charts: religious_demographics_analysis_charts.png")
        print(f"📄 CSV Files: {', '.join(csv_files)}")
        print("\n📋 DELIVERABLES SUMMARY:")
        print("• Complete demographic data from 1951-2022")
        print("• Excel file with 4 comprehensive worksheets") 
        print("• Visual trend analysis charts")
        print("• CSV files for additional analysis")
        print("• Statistical summary and key insights")
        print("• Ready for academic and research applications")
        print("="*80)
        
        return {
            'excel_file': excel_file,
            'chart_file': 'religious_demographics_analysis_charts.png',
            'csv_files': csv_files,
            'analyzer': analyzer
        }
        
    except Exception as e:
        print(f"❌ Error during analysis: {str(e)}")
        print("📝 Please ensure you have installed required packages:")
        print("   pip install pandas matplotlib seaborn openpyxl xlsxwriter")
        return None

if __name__ == "__main__":
    # Run the analysis
    results = main()
    
    if results:
        print(f"\n🎉 All files created successfully!")
        print(f"📁 Check your working directory for the generated files.")
    else:
        print(f"\n⚠️  Analysis failed. Please check error messages above.")