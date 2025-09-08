# -*- coding: utf-8 -*-
"""
Created on Mon Jul 21 10:41:37 2025

@author: muhuri
"""

# -*- coding: utf-8 -*-
"""
Enhanced Lab Data Extractor - Handles H/L flagged values and range values with Excel formatting
MAINTAINS ORIGINAL PDF ORDER
"""

import fitz
import pandas as pd
import re
import os
from datetime import datetime

def simple_extract_with_flags(pdf_path, excel_path):
    """Enhanced extraction that handles H/L flagged values like '102 H' and range values like '1-5'"""
    print("Starting ENHANCED extraction with H/L flag support...")
    
    if not os.path.exists(pdf_path):
        print("ERROR: PDF file not found")
        return None
    
    # Create output directory
    output_dir = os.path.dirname(excel_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    doc = fitz.open(pdf_path)
    extracted_data = []
    
    # Extended list of symbols that might represent triangle with ! 
    flag_symbols = ['!', '▲', '△', '⚠', '★', '*', '•', '◆', '♦', '▼', '⬥', '⚪', '🔺', '❗', '‼']
    
    for page_num, page in enumerate(doc, 1):
        print(f"Processing page {page_num}...")
        text = page.get_text()
        lines = text.split('\n')
        
        # METHOD 1: Original Reference Range pattern (enhanced to handle H/L flags and range values)
        for i, line in enumerate(lines):
            line = line.strip()
            
            # Look for "Reference Range:" lines
            if "Reference Range:" in line:
                print(f"  Found reference range: {line}")
                
                # Look backwards for the value
                value = None
                value_line_idx = None
                
                for j in range(i-1, max(i-3, -1), -1):
                    if j >= 0:
                        prev_line = lines[j].strip()
                        
                        # Check for H/L flagged values FIRST (higher priority)
                        if re.match(r'^\d+\.?\d*\s*[HL]$', prev_line):
                            value = prev_line
                            value_line_idx = j
                            print(f"    Found flagged value: {value}")
                            break
                        # Then check for plain numeric values
                        elif re.match(r'^\d+\.?\d*$', prev_line):
                            value = prev_line
                            value_line_idx = j
                            print(f"    Found value: {value}")
                            break
                        # Check for range values like "1-5"
                        elif re.match(r'^\d+\s*-\s*\d+$', prev_line):
                            value = prev_line
                            value_line_idx = j
                            print(f"    Found range value: {value}")
                            break
                        # Check if this is a text result
                        elif prev_line.upper() in ['NEGATIVE', 'POSITIVE', 'NORMAL', 'ABNORMAL', 'CLEAR', 'DETECTED', 'NOT DETECTED']:
                            value = prev_line
                            value_line_idx = j
                            print(f"    Found text value: {value}")
                            break
                
                if value and value_line_idx is not None:
                    # Look backwards for test name
                    test_name = None
                    test_name_line_idx = None
                    
                    for k in range(value_line_idx-1, max(value_line_idx-5, -1), -1):
                        if k >= 0:
                            test_line = lines[k].strip()
                            
                            # Skip empty lines and obvious non-test lines
                            if (not test_line or 
                                len(test_line) < 3 or 
                                test_line.startswith('Reference Range') or
                                re.match(r'^\d+\.?\d*$', test_line) or
                                re.match(r'^\d+\.?\d*\s*[HL]$', test_line) or
                                re.match(r'^\d+\s*-\s*\d+$', test_line)):
                                continue
                            
                            # This could be our test name
                            test_name = test_line
                            test_name_line_idx = k
                            print(f"    Found test name: {test_name}")
                            break
                    
                    if test_name and test_name_line_idx is not None:
                        # Parse reference range
                        ref_range = line.replace("Reference Range:", "").strip()
                        
                        # Handle H/L flagged values FIRST (highest priority)
                        if re.match(r'^\d+\.?\d*\s*[HL]$', value):
                            final_value = value  # Keep "102 H" as "102 H"
                            value_type = 'Flagged'
                            print(f"      → Flagged value: {final_value}")
                        # Then handle plain numeric values
                        elif re.match(r'^\d+\.?\d*$', value):
                            try:
                                final_value = float(value)
                                value_type = 'Numeric'
                                print(f"      → Numeric value: {final_value}")
                            except:
                                final_value = value
                                value_type = 'Text'
                        # Handle range values like "1-5"
                        elif re.match(r'^\d+\s*-\s*\d+$', value):
                            final_value = value
                            value_type = 'Range'
                            print(f"      → Range value: {final_value}")
                        # Default to text
                        else:
                            final_value = value
                            value_type = 'Text'
                        
                        # Check if this is one of our target tests with triangle symbols
                        is_triangle_test = test_name.upper() in [
                            'LDL-CHOLESTEROL', 'GLUCOSE', 'MCHC', 'HEMOGLOBIN A1C', 'HEMOGLOBIN A1c', 'HYALINE CAST'
                        ]
                        
                        result = {
                            'Test_Name': test_name,
                            'Value': final_value,
                            'Value_Type': value_type,
                            'Reference_Range': ref_range,
                            'Page': page_num,
                            'Line_Position': test_name_line_idx,  # KEY ADDITION: Track original position
                            'Has_Triangle_Symbol': is_triangle_test,
                            'Method': 'reference_range'
                        }
                        
                        extracted_data.append(result)
                        flag_note = " [TRIANGLE-!]" if is_triangle_test else ""
                        print(f"    ✓ EXTRACTED: {test_name} = {final_value}{flag_note} (Line {test_name_line_idx})")
        
        # METHOD 2: Look specifically for lines starting with triangle-! symbols
        print(f"  Searching for lines starting with triangle-! symbols...")
        for i, line in enumerate(lines):
            line = line.strip()
            
            # Check if line starts with any flag symbol
            starts_with_flag = any(line.startswith(symbol) for symbol in flag_symbols)
            
            if starts_with_flag and len(line) > 2:
                print(f"    FOUND TRIANGLE-! LINE: {line}")
                
                # Look ahead for a value in next few lines
                for j in range(i+1, min(i+8, len(lines))):
                    if j >= len(lines):
                        break
                    
                    value_line = lines[j].strip()
                    if not value_line:
                        continue
                    
                    # Check if this is a value (H/L flagged values get priority)
                    is_numeric_with_flag = re.match(r'^\d+\.?\d*\s*[HL]$', value_line)
                    is_numeric = re.match(r'^\d+\.?\d*$', value_line)
                    is_range = re.match(r'^\d+\s*-\s*\d+$', value_line)
                    is_text_result = value_line.upper() in ['NEGATIVE', 'POSITIVE', 'NORMAL', 'ABNORMAL', 'CLEAR', 'DETECTED', 'NOT DETECTED']
                    is_comparison = re.match(r'^[<>=]\s*\d+\.?\d*$', value_line)
                    
                    if is_numeric_with_flag or is_numeric or is_range or is_text_result or is_comparison:
                        print(f"      Found value for triangle-! test: {value_line}")
                        
                        # Look for reference range after the value
                        ref_range = ""
                        for k in range(j+1, min(j+4, len(lines))):
                            if k < len(lines) and "Reference Range:" in lines[k]:
                                ref_range = lines[k].replace("Reference Range:", "").strip()
                                break
                        
                        # Handle H/L flagged values FIRST (highest priority)
                        if is_numeric_with_flag:
                            final_value = value_line  # Keep "102 H" as "102 H"
                            value_type = 'Flagged'
                            print(f"        → Flagged value: {final_value}")
                        # Then handle plain numeric values
                        elif is_numeric:
                            try:
                                final_value = float(value_line)
                                value_type = 'Numeric'
                                print(f"        → Numeric value: {final_value}")
                            except:
                                final_value = value_line
                                value_type = 'Text'
                        elif is_range:
                            final_value = value_line
                            value_type = 'Range'
                            print(f"        → Range value: {final_value}")
                        # Default to text for other cases
                        else:
                            final_value = value_line
                            value_type = 'Text'
                        
                        # Only include if value is not missing/empty
                        if final_value and str(final_value).strip():
                            result = {
                                'Test_Name': line,
                                'Value': final_value,
                                'Value_Type': value_type,
                                'Reference_Range': ref_range if ref_range else None,
                                'Page': page_num,
                                'Line_Position': i,  # KEY ADDITION: Track original position
                                'Has_Triangle_Symbol': True,  # Lines starting with symbols have triangles
                                'Method': 'flag_symbol'
                            }
                            
                            extracted_data.append(result)
                            print(f"      ✓ CAPTURED TRIANGLE-! TEST: {line} = {final_value} ({value_type}) (Line {i})")
                            break
                        else:
                            print(f"      ✗ SKIPPED TRIANGLE-! TEST (empty value): {line}")
        
        # METHOD 3: Search for exact target test names (enhanced for H/L values and range values)
        target_test_names = [
            'LDL-CHOLESTEROL', 'LDL CHOLESTEROL', 'GLUCOSE', 'MCHC', 
            'HEMOGLOBIN A1C', 'HEMOGLOBIN A1c', 'LDL', 'HbA1c', 'A1C', 'HYALINE CAST'
        ]
        
        print(f"  Searching for exact target test names...")
        for i, line in enumerate(lines):
            line = line.strip()
            
            for target_name in target_test_names:
                if (line.upper() == target_name.upper() or 
                    line.upper().replace('-', ' ') == target_name.upper().replace('-', ' ') or
                    line.upper().replace(' ', '') == target_name.upper().replace(' ', '')):
                    
                    print(f"    FOUND EXACT TARGET: '{line}' matches '{target_name}'")
                    
                    # Skip if already captured (check by test name AND line position for better accuracy)
                    already_captured = False
                    for existing in extracted_data:
                        if (existing['Test_Name'] == line and 
                            existing['Page'] == page_num and 
                            abs(existing['Line_Position'] - i) <= 2):  # Allow small line difference
                            already_captured = True
                            break
                    
                    if already_captured:
                        print(f"      Already captured")
                        continue
                    
                    # Look ahead for a value
                    value_found = False
                    for j in range(i+1, min(i+8, len(lines))):
                        if j >= len(lines):
                            break
                        
                        value_line = lines[j].strip()
                        if not value_line:
                            continue
                        
                        print(f"      Checking line {j}: '{value_line}'")
                        
                        # Check if this is a value (H/L flagged values get priority)
                        is_numeric_with_flag = re.match(r'^\d+\.?\d*\s*[HL]$', value_line)
                        is_numeric = re.match(r'^\d+\.?\d*$', value_line)
                        is_range = re.match(r'^\d+\s*-\s*\d+$', value_line)
                        is_text_result = value_line.upper() in ['NEGATIVE', 'POSITIVE', 'NORMAL', 'ABNORMAL', 'CLEAR', 'DETECTED', 'NOT DETECTED']
                        is_comparison = re.match(r'^[<>=]\s*\d+\.?\d*$', value_line)
                        
                        if is_numeric_with_flag or is_numeric or is_range or is_text_result or is_comparison:
                            print(f"      ✓ Found value: {value_line}")
                            
                            # Look for reference range
                            ref_range = ""
                            for k in range(j+1, min(j+4, len(lines))):
                                if k < len(lines) and ("Reference Range:" in lines[k] or "Reference range:" in lines[k]):
                                    ref_range = lines[k].replace("Reference Range:", "").replace("Reference range:", "").strip()
                                    break
                            
                            # Handle H/L flagged values FIRST (highest priority)
                            if is_numeric_with_flag:
                                final_value = value_line  # Keep "102 H" as "102 H"
                                value_type = 'Flagged'
                                print(f"        → Flagged value: {final_value}")
                            # Then handle plain numeric values
                            elif is_numeric:
                                try:
                                    final_value = float(value_line)
                                    value_type = 'Numeric'
                                    print(f"        → Numeric value: {final_value}")
                                except:
                                    final_value = value_line
                                    value_type = 'Text'
                            elif is_range:
                                final_value = value_line
                                value_type = 'Range'
                                print(f"        → Range value: {final_value}")
                            # Default to text for other cases
                            else:
                                final_value = value_line
                                value_type = 'Text'
                            
                            # Check if this is one of our 5 target tests with triangle symbols
                            is_triangle_test = line.upper() in [
                                'LDL-CHOLESTEROL', 'GLUCOSE', 'MCHC', 'HEMOGLOBIN A1C', 'HEMOGLOBIN A1c', 'HYALINE CAST'
                            ]
                            
                            result = {
                                'Test_Name': line,
                                'Value': final_value,
                                'Value_Type': value_type,
                                'Reference_Range': ref_range if ref_range else None,
                                'Page': page_num,
                                'Line_Position': i,  # KEY ADDITION: Track original position
                                'Has_Triangle_Symbol': is_triangle_test,
                                'Method': 'target_name_search'
                            }
                            
                            extracted_data.append(result)
                            triangle_note = " [TRIANGLE-!]" if is_triangle_test else ""
                            print(f"      ✓ CAPTURED TARGET: {line} = {final_value}{triangle_note} (Line {i})")
                            value_found = True
                            break
                    
                    if not value_found:
                        print(f"      No value found for target: {line}")
                    
                    break
    
    doc.close()
    
    if extracted_data:
        df = pd.DataFrame(extracted_data)
        
        # KEY CHANGE: Sort by Page and Line_Position to maintain PDF order
        print(f"Sorting by original PDF order (Page, Line_Position)...")
        df = df.sort_values(['Page', 'Line_Position']).reset_index(drop=True)
        
        # Remove duplicates while preserving the new sorted order
        # Use a more sophisticated duplicate detection that considers line proximity
        df_filtered = []
        seen_tests = {}  # Changed to dict to store test info
        
        for _, row in df.iterrows():
            test_name_upper = row['Test_Name'].upper()
            current_page = row['Page']
            current_line = row['Line_Position']
            current_value = str(row['Value']) if row['Value'] is not None else ""
            
            # Check if we've seen this test before
            is_duplicate = False
            test_key = f"{test_name_upper}_{current_page}"
            
            if test_key in seen_tests:
                prev_line = seen_tests[test_key]['line']
                prev_value = seen_tests[test_key]['value']
                line_distance = abs(current_line - prev_line)
                
                # Consider it a duplicate if:
                # 1. Same test name on same page AND
                # 2. Lines are close together (within 5 lines) AND
                # 3. Values are the same (or both empty)
                if (line_distance <= 5 and 
                    (current_value == prev_value or 
                     (not current_value.strip() and not prev_value.strip()))):
                    is_duplicate = True
                    print(f"Skipped duplicate: {row['Test_Name']} (Line {current_line}) - too close to line {prev_line}")
                else:
                    # Different values or far apart lines - keep both but update the record
                    print(f"Kept similar test with different value/position: {row['Test_Name']} (Line {current_line}) vs previous (Line {prev_line})")
                    seen_tests[test_key] = {
                        'line': current_line,
                        'value': current_value
                    }
            else:
                # First time seeing this test
                seen_tests[test_key] = {
                    'line': current_line,
                    'value': current_value
                }
            
            if not is_duplicate:
                df_filtered.append(row)
                triangle_note = " [TRIANGLE-!]" if row['Has_Triangle_Symbol'] else ""
                print(f"Kept test: {row['Test_Name']} (Line {current_line}){triangle_note}")
        
        # Convert back to DataFrame
        df_filtered = pd.DataFrame(df_filtered)
        
        # Filter out missing values for non-triangle tests only
        df_final = df_filtered[
            (df_filtered['Value'].notna() & (df_filtered['Value'].astype(str).str.strip() != '')) |
            (df_filtered['Has_Triangle_Symbol'] == True)
        ].copy()
        
        # Reset index to maintain sorted order
        df_final = df_final.reset_index(drop=True)
        
        print(f"Final dataset maintains original PDF order...")
        
        df_final['Extraction_Date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Select only the required columns
        columns_to_keep = ['Test_Name', 'Value', 'Reference_Range', 'Extraction_Date']
        df_output = df_final[columns_to_keep].copy()
        
        print(f"Keeping only columns: {', '.join(columns_to_keep)}")
        
        # Save to Excel first
        df_output.to_excel(excel_path, index=False)
        
        # Apply formatting to H/L flagged values
        print(f"Applying formatting to H/L flagged values...")
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, Alignment
            
            # Load the workbook and worksheet
            wb = load_workbook(excel_path)
            ws = wb.active
            
            # Define styles for flagged values and general alignment
            red_font = Font(color='FF0000', bold=True)  # Red color
            right_align = Alignment(horizontal='right')
            
            # Find the Value and Reference_Range columns
            value_col_idx = None
            ref_range_col_idx = None
            for col_idx, cell in enumerate(ws[1], 1):  # Header row
                if cell.value == 'Value':
                    value_col_idx = col_idx
                elif cell.value == 'Reference_Range':
                    ref_range_col_idx = col_idx
            
            # Right-align Value column
            if value_col_idx:
                print(f"  Found Value column at index {value_col_idx}")
                
                # Apply right-alignment to ALL values in the Value column and special formatting for flagged values
                flagged_count = 0
                total_values = 0
                chol_hdlc_formatted = 0
                
                for row_idx in range(2, ws.max_row + 1):  # Skip header row
                    cell = ws.cell(row=row_idx, column=value_col_idx)
                    
                    # Right-align ALL values in the Value column
                    cell.alignment = right_align
                    total_values += 1
                    
                    # Get the test name from the same row (assuming Test_Name is in column A)
                    test_name_cell = ws.cell(row=row_idx, column=1)  # Column A
                    test_name = str(test_name_cell.value) if test_name_cell.value else ""
                    
                    # Special formatting for CHOL/HDLC RATIO - add one decimal place
                    if (test_name.upper() == "CHOL/HDLC RATIO" or 
                        "CHOL/HDLC" in test_name.upper() or
                        "CHOLESTEROL/HDL" in test_name.upper()):
                        
                        if cell.value is not None:
                            try:
                                # Convert to float and format with 1 decimal place
                                numeric_value = float(str(cell.value).replace(' H', '').replace(' L', '').strip())
                                
                                # Check if original had H or L flag
                                original_str = str(cell.value)
                                has_h_flag = original_str.endswith(' H') or original_str.endswith('H')
                                has_l_flag = original_str.endswith(' L') or original_str.endswith('L')
                                
                                # Format with 1 decimal place
                                formatted_value = f"{numeric_value:.1f}"
                                
                                # Add back the H or L flag if it existed
                                if has_h_flag:
                                    formatted_value += " H"
                                elif has_l_flag:
                                    formatted_value += " L"
                                
                                cell.value = formatted_value
                                chol_hdlc_formatted += 1
                                print(f"    Formatted CHOL/HDLC RATIO: {original_str} → {formatted_value}")
                                
                                # Apply red formatting if it has H or L flag
                                if has_h_flag or has_l_flag:
                                    cell.font = red_font
                                    flagged_count += 1
                                    
                            except (ValueError, TypeError):
                                print(f"    Could not format CHOL/HDLC RATIO value: {cell.value}")
                    
                    # Regular flagged value formatting for other tests
                    elif cell.value and isinstance(cell.value, str):
                        cell_str = str(cell.value)
                        
                        # Check if the value contains H or L flag at the end
                        if cell_str.endswith(' H') or cell_str.endswith(' L') or cell_str.endswith('H') or cell_str.endswith('L'):
                            cell.font = red_font
                            flagged_count += 1
                            print(f"    Formatted flagged value: {cell.value}")
                
                print(f"  Applied right-alignment to {total_values} values in Value column")
                print(f"  Applied red formatting to {flagged_count} flagged values")
                if chol_hdlc_formatted > 0:
                    print(f"  Added 1 decimal place to {chol_hdlc_formatted} CHOL/HDLC RATIO values")
            else:
                print("  Warning: Could not find Value column for formatting")
            
            # Right-align Reference_Range column
            if ref_range_col_idx:
                print(f"  Found Reference_Range column at index {ref_range_col_idx}")
                
                ref_range_count = 0
                for row_idx in range(2, ws.max_row + 1):  # Skip header row
                    cell = ws.cell(row=row_idx, column=ref_range_col_idx)
                    if cell.value:  # Only align if there's content
                        cell.alignment = right_align
                        ref_range_count += 1
                
                print(f"  Applied right-alignment to {ref_range_count} reference ranges")
            else:
                print("  Warning: Could not find Reference_Range column")
                
            # Auto-adjust column widths for better readability
            print(f"  Auto-sizing column widths...")
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set minimum width and add padding
                adjusted_width = min(max_length + 2, 50)  # Max width of 50, min padding of 2
                if adjusted_width < 12:  # Minimum column width
                    adjusted_width = 12
                
                ws.column_dimensions[column_letter].width = adjusted_width
                print(f"    Column {column_letter}: width set to {adjusted_width}")
            
            # Save the formatted workbook
            wb.save(excel_path)
            print(f"  Formatting and column sizing saved to Excel file")
                
        except ImportError:
            print("  Warning: openpyxl not available. Install with: pip install openpyxl")
            print("  H/L values saved but not formatted")
        except Exception as e:
            print(f"  Warning: Could not apply formatting: {e}")
            print("  H/L values saved but not formatted")
        
        print(f"\nResults saved to: {excel_path}")
        
        # Show summary
        ref_range_tests = df_final[df_final['Method'] == 'reference_range']
        target_name_tests = df_final[df_final['Method'] == 'target_name_search']
        
        print(f"\nSUMMARY:")
        print(f"  Total tests: {len(df_final)}")
        print(f"  Reference Range method: {len(ref_range_tests)}")
        print(f"  Tests with triangle symbols: {df_final['Has_Triangle_Symbol'].sum()}")
        
        # Show target tests found
        target_found = df_final[df_final['Has_Triangle_Symbol'] == True]
        if len(target_found) > 0:
            print(f"\n✅ TESTS WITH TRIANGLE-! SYMBOLS:")
            print("-" * 70)
            for i, (_, row) in enumerate(target_found.iterrows(), 1):
                value_display = row['Value'] if row['Value'] is not None else "MISSING"
                print(f"  {i}. '{row['Test_Name']}': {value_display} ({row['Value_Type']}) - Line {row['Line_Position']}")
        else:
            print(f"\n❌ NO TRIANGLE-! TESTS FOUND")
        
        # Show all tests for verification WITH LINE POSITIONS
        print(f"\nALL EXTRACTED TESTS (IN PDF ORDER):")
        for i, (_, row) in enumerate(df_final.iterrows(), 1):
            triangle_note = " [TRIANGLE-!]" if row['Has_Triangle_Symbol'] else ""
            value_display = row['Value'] if row['Value'] is not None else "MISSING"
            print(f"  {i:2d}. '{row['Test_Name']}': {value_display}{triangle_note} (Line {row['Line_Position']})")
        
        return df_final
    else:
        print("No data extracted")
        return None

if __name__ == "__main__":
    print("="*70)
    print("ENHANCED LAB DATA EXTRACTOR - H/L FLAG SUPPORT WITH PDF ORDER PRESERVATION")
    print("="*70)
    
    pdf_file = r'C:\Data\LabReport_2025.pdf'
    excel_file = r'C:\Data2\LabReport_2025_enhanced.xlsx'
    
    result = simple_extract_with_flags(pdf_file, excel_file)
    
    if result is not None:
        print(f"\nEXTRACTION COMPLETE!")
        print(f"Found {len(result)} total lab tests")
        
        target_count = len(result[result['Has_Triangle_Symbol'] == True])
        if target_count > 0:
            print(f"✅ Found {target_count} tests with triangle-! symbols")
        else:
            print(f"❌ No triangle-! tests found")
        
        print(f"Check output: {excel_file}")
        print(f"Tests are now ordered exactly as they appear in the PDF!")
    else:
        print("No lab tests found")
    
    print("="*70)