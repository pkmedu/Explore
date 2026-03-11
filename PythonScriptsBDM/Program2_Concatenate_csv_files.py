# -*- coding: utf-8 -*-
"""
Created on Mon May  5 12:32:20 2025
@author: muhuri
"""

import pandas as pd
import glob
import os
import logging
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(message)s')

# Set paths
combined_csv_path = 'C:/Explore/PythonScriptsBDM/dashboard.csv'
final_excel_path = 'C:/Explore/PythonScriptsBDM/dashboard.xlsx'
final_html_path = 'C:/Explore/PythonScriptsBDM/dashboard.html'

# Ensure output directory exists
os.makedirs(os.path.dirname(final_excel_path), exist_ok=True)

# Find all CSV files in the folder
folder_path = 'c:/Data/CSV_updated'  
csv_files = glob.glob(os.path.join(folder_path, '*.csv'))
if not csv_files:
    raise FileNotFoundError(f"No CSV files found in {folder_path}")

# Read and combine CSVs
df_list = [pd.read_csv(file) for file in csv_files]
combined_df = pd.concat(df_list, ignore_index=True)

# Save combined data as CSV (optional)
combined_df.to_csv(combined_csv_path, index=False)

# Reload to ensure consistency
df = pd.read_csv(combined_csv_path)

# Count rows before and after deduplication
initial_count = len(df)

# Ensure 'Link' column exists
if 'Link' not in df.columns:
    raise ValueError("Column 'Link' not found in the CSV files.")

# Drop duplicates based on 'Link'
df_unique = df.drop_duplicates(subset='Link', keep='first')
Excl_dups_count = len(df_unique)
duplicates_removed = initial_count - Excl_dups_count

# Sort by 'Link'
df_unique_sorted = df_unique.sort_values(by='Link')

# Get count before filtering
before_filter_count = len(df_unique_sorted)

exclude_terms = ['#aiart', '#iran', 'CANADIAN', 'Havildar', 'Izak', 'MUSIC', 'Bhutan', 'Malaysia',
                 '#thinkbasicfolks',  'fifth grader', 'Syria', 'MP Police Constable', 'UP POLICE',
                 'ဓမ', 'Song', 'Assam', 'BJP Rips','LGBTQ', 'Class 12th', 'Laxmikant', 'Iran',
                 'Internet', '1st paper', '6th', 'နေ့စဉ်', 'Culture', '#IqraHassan',
                 'filippino','老王来了', 'class 12', 'ه الرخيصه', 'মাইলস্টোন ' , 'Over Bihar',
                 'Electoral Integrity', 'Ancient Traditions', 'কেমন বিচার' , ' Sachin Academy',
                 'Iqra Hasa', 'Aadhaar', 'Election Commission', 'Islamabad High Court',
                                 ' পরিযায়ী ',   ' Haat', 'Philippines',  'CAA',
                                 ' হারাম ', ' REALLY RULE THE COUNTRY', 'মমতার চাপের কাছে'  , '【德媒】',
                'Anti - Muslim?', 'ভক্তিবেদান্ত স্বামী প্রভুপাদ', ' আটক' , ' 德媒','ညစဥ်ဖွင့်၍',
                 'မိမိစိတ်ကလေး', 'သစ္စာရွေစည်ဆရာတော်',
                 'သစ္စာရွေစည်ဆရာတော်', 'မိမိစိတ်ကလေး အနားရဖို့',
                 'ညစဥ်ဖွင့်၍', 'ညစဥ်ဖွင့်၍']

exclude_pattern = '|'.join(exclude_terms)
keep_mask = ~df_unique_sorted['Title'].astype(str).str.contains(exclude_pattern, case=False, na=False)
df_unique_after_filter = df_unique_sorted[keep_mask].copy()

after_filter_count = len(df_unique_after_filter)
exclusion_count = before_filter_count - after_filter_count

df_unique_after_filter = df_unique_after_filter.sort_values(
    by=['Hashtag', 'Publication Date'],
    ascending=[True, False]
)

# -------------------------
# Save Excel with hyperlinks
# -------------------------
wb = Workbook()
ws = wb.active
ws.title = 'Links'

# Write header row
ws.append(list(df_unique_after_filter.columns))

# Write data rows with hyperlinks
for row in df_unique_after_filter.itertuples(index=False):
    ws_row = ws.max_row + 1
    for col_idx, value in enumerate(row):
        cell = ws.cell(row=ws_row, column=col_idx + 1)
        col_name = df_unique_after_filter.columns[col_idx]
        if col_name == 'Link' and pd.notnull(value):
            cell.value = value
            cell.hyperlink = value
            cell.font = Font(color='0000FF', underline='single')  # Blue underlined
        else:
            cell.value = value

# Adjust column widths
for col_idx, column in enumerate(df_unique_after_filter.columns, 1):
    max_length = max((len(str(cell)) for cell in df_unique_after_filter[column]), default=0)
    adjusted_width = min(max_length + 2, 100)
    ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

# Save Excel file
wb.save(final_excel_path)

# -------------------------
# Save HTML file with clickable hyperlinks
# -------------------------
# Convert 'Link' column values into HTML anchor tags
df_html = df_unique_after_filter.copy()
if 'Link' in df_html.columns:
    df_html['Link'] = df_html['Link'].apply(
        lambda url: f'<a href="{url}" target="_blank">{url}</a>' if pd.notnull(url) else ''
    )

# Save as HTML table
table = df_html.to_html(escape=False, index=False)

with open(final_html_path, 'w', encoding='utf-8') as file:
    file.writelines('<meta charset="utf-8">\n')
    file.write(table)

# -------------------------
# Log summary
# -------------------------
logging.info(f"Total rows before deduplication: {initial_count}")
logging.info(f"Total rows after deduplication: {Excl_dups_count}")
logging.info(f"Duplicates removed: {duplicates_removed}")
logging.info(f"Total rows after filtering out: {after_filter_count}")
logging.info(f"Excluded rows: {exclusion_count}")
logging.info(f"Excel file saved to: {final_excel_path}")
logging.info(f"HTML table saved to: {final_html_path}")
